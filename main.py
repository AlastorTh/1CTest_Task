import openpyxl
import os
import sys
from classes import Sells, ReceiptLine
from mapping import sells_page, sells_id, sells_left, sells_price, receipt_page


print("Укажите путь в папку, в которой будет находиться исходный файл")
print("Пример: C:\\Users\\Касаткины\\Desktop\\Exceltest", "-двойные \\")
folder_path = input()
os.chdir(folder_path)
# если нет желания вводить путь каждый раз, можно ввести его вручную 1 раз
# os.chdir("C:\\Users\\Касаткины\\Desktop\\Exceltest")
print("Введите имя excel файла с данными о товарах в формате filename.xlsx")
print("Формат файла входных данных: страница Продажи с товарами, пустая страница Счёт")
file_path = input()
# аналогично можно ввести путь к файлу 1 раз
# wb = openpyxl.load_workbook("example.xlsx")
wb = openpyxl.load_workbook(file_path)


def read_sells(sheet):  # считывание данных для продажи
    products = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        product = Sells(title=row[sells_id], left=row[sells_left], price=row[sells_price])
        products.append(product)

    return products


def seek_title(titles, split): # поиск наименования в таблице

    if split[0] in titles:
        index = titles.index(split[0])
        return True
    else:
        return False


def create_receipt(split, selling_price, receipt_list):  # составление строки счёта

    selling = ReceiptLine(title=split[0], amount=split[1], price=selling_price, bulk_price=selling_price*int(split[1]))
    receipt_list.append(selling)
    return receipt_list


def sell_buy_product(products_list, receipt_flag):  # продажа и покупка товаров

    print("Введите выполняемую операцию(закупка/продажа)")
    print("Для завершения работы с покупкой/продажей наберите завершить")
    titles = []
    receipt_list = []

    for i in range(len(products_list)):  # создание списка имён
        titles.append(products_list[i].title)

    while True:
        call = input()
        if call == "продажа":
            print("Введите наименование и количество товара, разделитель - точка. Пример: Мешок семян.1")
            good = input()
            split = good.split(".")
            if seek_title(titles, split):
                index = titles.index(split[0])
                if int(split[1]) > 0:
                    products_list[index].left -= int(split[1])
                    selling_price = float(products_list[index].price)
                    receipt = create_receipt(split, selling_price, receipt_list)
                    receipt_flag = 1
                    print(split[0], "продано")
                else:
                    print("ОШИБКА! Введено отрицательное число для продажи. Вы можете продолжить ввод")
                    continue
                if products_list[index].left < 0:
                    print("ОШИБКА! Было продано число товара, большее остатка")
                    sys.exit()
            else:
                print("ОШИБКА! Данного наименования нет в продаже ")
                continue
        elif call == "закупка":
            print("Введите наименование и количество товара")
            good = input()
            split = good.split(".")
            if seek_title(titles, split):
                index = titles.index(split[0])
                if int(split[1]) > 0:
                    products_list[index].left += int(split[1])
                    print(split[0], "поступило на склад")
                else:
                    print("ОШИБКА! Введено отрицательное число для закупки")
                    continue
            else:
                print("Добавлено новое наименование! Укажите цену")
                new_price = float(input())
                if new_price > 0:
                    products_list.append(Sells(title=split[0], left=int(split[1]), price=new_price))
                    titles.append(split[0])
                    print(split[0], "поступило на склад")
                else:
                    print("ОШИБКА! Указано отрицательное значение стоимости товара")
                    sys.exit()
        elif call == "завершить":
            if receipt_flag:
                return receipt
            else:
                return
        else:
            print("Неизвестная команда! Введите команды закупка/продажа/завершить")
            continue


action = True
products = read_sells(wb.worksheets[sells_page])
receipt_flag = 0


while action:
    print("\nВведите команду ""продажи"" для перехода к работе с изменением числа товаров на складе\n")
    print("Введите ""сохранить"" для перехода к экспорту в excel")
    function_command = input()
    if function_command == "продажи":
        receipt = sell_buy_product(products, receipt_flag)
        overall_price = 0
        wb.remove(wb["Продажи"])
        wb.remove(wb["Счёт"])
        wb.create_sheet("Продажи", 0)
        wb.create_sheet("Счёт", 1)
        if receipt_flag:
            for i in receipt:
                overall_price += i.bulk_price
                print(i.title, i.amount, i.price, i.bulk_price)

    elif function_command == "сохранить":
        origin_set = ["Наименование", "Остаток", "Цена"]
        wb.worksheets[sells_page].append(origin_set)
        origin_set = ["Наименование", "Количество", "Цена", "Стоимость"]
        wb.worksheets[receipt_page].append(origin_set)
        for product in products:
            data = [product.title, product.left, product.price]
            wb.worksheets[sells_page].append(data)
        if receipt_flag == 0:  # список пуст
            print("Счёт пуст!")
            break
        else:
            sum_counter = 0
            for receipt_row in receipt:
                receipt_data = [receipt_row.title, receipt_row.amount, receipt_row.price, receipt_row.bulk_price]
                sum_counter += receipt_row.bulk_price
                wb.worksheets[receipt_page].append(receipt_data)
            ending = ["", "", "Итого", sum_counter]
            wb.worksheets[receipt_page].append(ending)
            break



wb.save("mod.xlsx")
